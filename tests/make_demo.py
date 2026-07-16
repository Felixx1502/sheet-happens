#!/usr/bin/env python3
"""Generate DemoCo — a small synthetic 3-statement model used for the public
demo map and screenshots. No real company data."""
import os
from openpyxl import Workbook


def build(path):
    wb = Workbook()
    # ---------------- Assumptions ----------------
    ws = wb.active
    ws.title = 'Assumptions'
    rows = [
        ('DEMOCO — ASSUMPTIONS',), (),
        ('REVENUE DRIVERS',),
        ('Units sold Y1', 1000), ('Unit growth %', 0.12), ('Price / unit', 25.0),
        ('Price growth %', 0.03), (),
        ('COST DRIVERS',),
        ('COGS % of revenue', 0.42), ('Opex fixed', 4000), ('Opex % of revenue', 0.08), (),
        ('CAPEX & DEPRECIATION',),
        ('Capex % of revenue', 0.05), ('Depreciation years', 5), (),
        ('WORKING CAPITAL & TAX',),
        ('DSO days', 45), ('DPO days', 30), ('Tax rate', 0.2), (),
        ('DEBT',),
        ('Opening debt', 8000), ('Interest rate', 0.07), ('Annual repayment', 1000),
    ]
    for r in rows:
        ws.append(list(r) if r else [None])

    # ---------------- Revenue Build ----------------
    ws = wb.create_sheet('Revenue Build')
    ws.append(['DEMOCO — REVENUE BUILD', None, None, None, None, None])
    ws.append(['Year', 1, 2, 3, 4, 5])
    ws.append(['Units'] + [None] * 5)
    ws.append(['Price'] + [None] * 5)
    ws.append(['Revenue'] + [None] * 5)
    ws['B3'] = '=Assumptions!B4'
    ws['B4'] = '=Assumptions!B6'
    for c in 'CDEF':
        prev = chr(ord(c) - 1)
        ws[f'{c}3'] = f'={prev}3*(1+Assumptions!$B$5)'
        ws[f'{c}4'] = f'={prev}4*(1+Assumptions!$B$7)'
    for c in 'BCDEF':
        ws[f'{c}5'] = f'={c}3*{c}4'

    # ---------------- Cost Build ----------------
    ws = wb.create_sheet('Cost Build')
    ws.append(['DEMOCO — COST BUILD', None, None, None, None, None])
    ws.append(['Year', 1, 2, 3, 4, 5])
    ws.append(['COGS'] + [None] * 5)
    ws.append(['Opex'] + [None] * 5)
    ws.append(['Capex'] + [None] * 5)
    ws.append(['Depreciation'] + [None] * 5)
    for i, c in enumerate('BCDEF'):
        ws[f'{c}3'] = f"='Revenue Build'!{c}5*Assumptions!$B$10"
        ws[f'{c}4'] = f"=Assumptions!$B$11+'Revenue Build'!{c}5*Assumptions!$B$12"
        ws[f'{c}5'] = f"='Revenue Build'!{c}5*Assumptions!$B$15"
        # simple: depreciate cumulative capex straight-line
        ws[f'{c}6'] = f'=SUM($B$5:{c}5)/Assumptions!$B$16'

    # ---------------- Debt Schedule ----------------
    ws = wb.create_sheet('Debt Schedule')
    ws.append(['DEMOCO — DEBT SCHEDULE', None, None, None, None, None])
    ws.append(['Year', 1, 2, 3, 4, 5])
    ws.append(['Opening debt'] + [None] * 5)
    ws.append(['Repayment'] + [None] * 5)
    ws.append(['Closing debt'] + [None] * 5)
    ws.append(['Interest'] + [None] * 5)
    ws['B3'] = '=Assumptions!B22'
    for c in 'CDEF':
        ws[f'{c}3'] = f'={chr(ord(c)-1)}5'
    for c in 'BCDEF':
        ws[f'{c}4'] = f'=MIN(Assumptions!$B$24,{c}3)'
        ws[f'{c}5'] = f'={c}3-{c}4'
        ws[f'{c}6'] = f'={c}3*Assumptions!$B$23'

    # ---------------- P&L ----------------
    ws = wb.create_sheet('P&L')
    ws.append(['DEMOCO — P&L', None, None, None, None, None])
    ws.append(['Year', 1, 2, 3, 4, 5])
    labels = ['Revenue', 'COGS', 'Gross profit', 'Opex', 'EBITDA',
              'Depreciation', 'EBIT', 'Interest', 'PBT', 'Tax', 'Net income']
    for lab in labels:
        ws.append([lab] + [None] * 5)
    for c in 'BCDEF':
        ws[f'{c}3'] = f"='Revenue Build'!{c}5"
        ws[f'{c}4'] = f"='Cost Build'!{c}3"
        ws[f'{c}5'] = f'={c}3-{c}4'
        ws[f'{c}6'] = f"='Cost Build'!{c}4"
        ws[f'{c}7'] = f'={c}5-{c}6'
        ws[f'{c}8'] = f"='Cost Build'!{c}6"
        ws[f'{c}9'] = f'={c}7-{c}8'
        ws[f'{c}10'] = f"='Debt Schedule'!{c}6"
        ws[f'{c}11'] = f'={c}9-{c}10'
        ws[f'{c}12'] = f'=MAX({c}11,0)*Assumptions!$B$19'
        ws[f'{c}13'] = f'={c}11-{c}12'

    # ---------------- Cash Flow ----------------
    ws = wb.create_sheet('Cash Flow')
    ws.append(['DEMOCO — CASH FLOW', None, None, None, None, None])
    ws.append(['Year', 1, 2, 3, 4, 5])
    for lab in ['Net income', 'Add back D&A', 'Change in WC', 'Capex',
                'Debt repayment', 'Net cash flow', 'Cash balance']:
        ws.append([lab] + [None] * 5)
    for i, c in enumerate('BCDEF'):
        ws[f'{c}3'] = f"='P&L'!{c}13"
        ws[f'{c}4'] = f"='Cost Build'!{c}6"
        ws[f'{c}5'] = f"=-('P&L'!{c}3*Assumptions!$B$18/365-'P&L'!{c}4*Assumptions!$B$19/365)"
        ws[f'{c}6'] = f"=-'Cost Build'!{c}5"
        ws[f'{c}7'] = f"=-'Debt Schedule'!{c}4"
        ws[f'{c}8'] = f'=SUM({c}3:{c}7)'
        ws[f'{c}9'] = (f'={c}8' if c == 'B' else f'={chr(ord(c)-1)}9+{c}8')

    # ---------------- Returns (with a deliberate hardcode) ----------------
    ws = wb.create_sheet('Returns')
    ws.append(['DEMOCO — RETURNS'])
    ws.append(['Exit multiple (hardcoded!)', None])
    ws.append(['Exit EBITDA', None])
    ws.append(['Enterprise value', None])
    ws.append(['Net debt at exit', None])
    ws.append(['Equity value', None])
    ws['B3'] = "='P&L'!F7*8.5"
    ws['B4'] = '=B3'
    ws['B5'] = "='Debt Schedule'!F5-'Cash Flow'!F9"
    ws['B6'] = '=B4-B5'

    wb.save(path)
    print('demo model written:', path)


if __name__ == '__main__':
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'demo',
                       'DemoCo_model.xlsx')
    os.makedirs(os.path.dirname(out), exist_ok=True)
    build(out)
